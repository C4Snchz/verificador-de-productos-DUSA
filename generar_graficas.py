#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de Gráficas y Tablas - Verificador DUSA
==================================================
Crea visualizaciones y un Excel con tablas dinámicas
"""

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Para guardar sin mostrar
import seaborn as sns
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

# Configuración
ARCHIVO_ENTRADA = "/Users/carlossanchez/Downloads/resultado_dusa_20260301_0414 (1).xlsx"
CARPETA_SALIDA = "/Users/carlossanchez/Downloads/analisis_dusa"

# Crear carpeta de salida
os.makedirs(CARPETA_SALIDA, exist_ok=True)

# Leer datos
print("📊 Leyendo archivo de resultados...")
df = pd.read_excel(ARCHIVO_ENTRADA)

# Limpiar columna Estado DUSA (quitar emojis para análisis)
def limpiar_estado(estado):
    if pd.isna(estado):
        return 'Sin datos'
    estado = str(estado)
    if 'Disponible' in estado:
        return 'Disponible'
    elif 'Faltante' in estado or 'Faltando' in estado:
        return 'Faltante'
    elif 'Consultar' in estado:
        return 'Consultar'
    elif 'No encontrado' in estado:
        return 'No encontrado'
    return estado

df['Estado_Limpio'] = df['Estado DUSA'].apply(limpiar_estado)

# Limpiar columna Accion
df['Accion_Limpia'] = df['Accion'].fillna('Sin acción')

print(f"   Total de productos: {len(df)}")

# ============================================
# 1. GRÁFICO DE TORTA - ESTADOS DUSA
# ============================================
print("\n📈 Generando gráfico de estados DUSA...")

estados_count = df['Estado_Limpio'].value_counts()
colores_estados = {
    'Disponible': '#28a745',
    'Faltante': '#dc3545',
    'Consultar': '#ffc107',
    'No encontrado': '#6c757d',
    'Sin datos': '#adb5bd'
}
colores = [colores_estados.get(e, '#17a2b8') for e in estados_count.index]

fig, ax = plt.subplots(figsize=(10, 8))
wedges, texts, autotexts = ax.pie(
    estados_count.values, 
    labels=estados_count.index,
    autopct=lambda pct: f'{pct:.1f}%\n({int(pct/100*sum(estados_count.values))})',
    colors=colores,
    explode=[0.02] * len(estados_count),
    shadow=True,
    startangle=90
)
ax.set_title('Distribución de Estados en DUSA\n', fontsize=16, fontweight='bold')
plt.setp(autotexts, size=10, weight='bold')
plt.tight_layout()
plt.savefig(f'{CARPETA_SALIDA}/grafico_estados_dusa.png', dpi=150, bbox_inches='tight')
plt.close()
print("   ✅ Guardado: grafico_estados_dusa.png")

# ============================================
# 2. GRÁFICO DE BARRAS - ESTADOS DUSA
# ============================================
fig, ax = plt.subplots(figsize=(12, 6))
bars = ax.bar(estados_count.index, estados_count.values, color=colores, edgecolor='black')
ax.set_title('Cantidad de Productos por Estado en DUSA', fontsize=14, fontweight='bold')
ax.set_xlabel('Estado')
ax.set_ylabel('Cantidad de Productos')

# Añadir valores encima de las barras
for bar, val in zip(bars, estados_count.values):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 20, 
            f'{val:,}', ha='center', va='bottom', fontweight='bold', fontsize=11)

plt.tight_layout()
plt.savefig(f'{CARPETA_SALIDA}/grafico_barras_estados.png', dpi=150, bbox_inches='tight')
plt.close()
print("   ✅ Guardado: grafico_barras_estados.png")

# ============================================
# 3. GRÁFICO DE TORTA - ACCIONES TOMADAS
# ============================================
print("\n📈 Generando gráfico de acciones tomadas...")

acciones_count = df['Accion_Limpia'].value_counts()
colores_acciones = {
    'Activado': '#28a745',
    'No activado': '#dc3545',
    'Incumplimiento': '#fd7e14',
    'Diferida': '#ffc107',
    'Sin acción': '#6c757d',
    'Hay que crear una publicacion nueva': '#17a2b8'
}
colores_acc = [colores_acciones.get(a, '#6c757d') for a in acciones_count.index]

fig, ax = plt.subplots(figsize=(10, 8))
wedges, texts, autotexts = ax.pie(
    acciones_count.values, 
    labels=[f'{l}\n({v})' for l, v in zip(acciones_count.index, acciones_count.values)],
    autopct='%1.1f%%',
    colors=colores_acc,
    explode=[0.02] * len(acciones_count),
    shadow=True,
    startangle=90
)
ax.set_title('Acciones Tomadas en los Productos\n', fontsize=16, fontweight='bold')
plt.setp(autotexts, size=9, weight='bold')
plt.tight_layout()
plt.savefig(f'{CARPETA_SALIDA}/grafico_acciones.png', dpi=150, bbox_inches='tight')
plt.close()
print("   ✅ Guardado: grafico_acciones.png")

# ============================================
# 4. GRÁFICO DE BARRAS - ACCIONES POR ESTADO
# ============================================
print("\n📈 Generando gráfico de acciones por estado...")

# Solo productos con acciones (no "Sin acción")
df_con_accion = df[df['Accion_Limpia'] != 'Sin acción']

if len(df_con_accion) > 0:
    pivot = pd.crosstab(df_con_accion['Estado_Limpio'], df_con_accion['Accion_Limpia'])
    
    fig, ax = plt.subplots(figsize=(14, 7))
    pivot.plot(kind='bar', ax=ax, colormap='Set2', edgecolor='black')
    ax.set_title('Acciones Tomadas por Estado de Producto', fontsize=14, fontweight='bold')
    ax.set_xlabel('Estado en DUSA')
    ax.set_ylabel('Cantidad')
    ax.legend(title='Acción', bbox_to_anchor=(1.02, 1), loc='upper left')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig(f'{CARPETA_SALIDA}/grafico_acciones_por_estado.png', dpi=150, bbox_inches='tight')
    plt.close()
    print("   ✅ Guardado: grafico_acciones_por_estado.png")

# ============================================
# 5. CREAR EXCEL CON TABLAS DINÁMICAS
# ============================================
print("\n📊 Generando Excel con tablas dinámicas...")

wb = Workbook()

# Estilos
header_fill = PatternFill(start_color="1e3c72", end_color="1e3c72", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
disponible_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
faltante_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
consultar_fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ----- HOJA 1: RESUMEN GENERAL -----
ws_resumen = wb.active
ws_resumen.title = "Resumen General"

# Título
ws_resumen['A1'] = "ANÁLISIS DE VERIFICACIÓN DUSA"
ws_resumen['A1'].font = Font(bold=True, size=18, color="1e3c72")
ws_resumen.merge_cells('A1:D1')

ws_resumen['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
ws_resumen['A2'].font = Font(italic=True, color="666666")

# Resumen Estados
ws_resumen['A4'] = "RESUMEN POR ESTADO EN DUSA"
ws_resumen['A4'].font = Font(bold=True, size=14)
ws_resumen.merge_cells('A4:C4')

headers = ['Estado', 'Cantidad', 'Porcentaje']
for col, header in enumerate(headers, 1):
    cell = ws_resumen.cell(row=5, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

total = len(df)
for row_idx, (estado, count) in enumerate(estados_count.items(), 6):
    ws_resumen.cell(row=row_idx, column=1, value=estado).border = thin_border
    ws_resumen.cell(row=row_idx, column=2, value=count).border = thin_border
    ws_resumen.cell(row=row_idx, column=3, value=f"{count/total*100:.1f}%").border = thin_border
    
    # Colorear según estado
    if 'Disponible' in estado:
        for col in range(1, 4):
            ws_resumen.cell(row=row_idx, column=col).fill = disponible_fill
    elif 'Faltante' in estado:
        for col in range(1, 4):
            ws_resumen.cell(row=row_idx, column=col).fill = faltante_fill
    elif 'Consultar' in estado:
        for col in range(1, 4):
            ws_resumen.cell(row=row_idx, column=col).fill = consultar_fill

# Total
row_total = 6 + len(estados_count)
ws_resumen.cell(row=row_total, column=1, value="TOTAL").font = Font(bold=True)
ws_resumen.cell(row=row_total, column=2, value=total).font = Font(bold=True)
ws_resumen.cell(row=row_total, column=3, value="100%").font = Font(bold=True)

# Resumen Acciones
ws_resumen['A13'] = "RESUMEN POR ACCIÓN TOMADA"
ws_resumen['A13'].font = Font(bold=True, size=14)
ws_resumen.merge_cells('A13:C13')

for col, header in enumerate(headers, 1):
    cell = ws_resumen.cell(row=14, column=col, value=header.replace('Estado', 'Acción'))
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for row_idx, (accion, count) in enumerate(acciones_count.items(), 15):
    ws_resumen.cell(row=row_idx, column=1, value=accion).border = thin_border
    ws_resumen.cell(row=row_idx, column=2, value=count).border = thin_border
    ws_resumen.cell(row=row_idx, column=3, value=f"{count/total*100:.1f}%").border = thin_border
    
    if accion == 'Activado':
        for col in range(1, 4):
            ws_resumen.cell(row=row_idx, column=col).fill = disponible_fill

# Ajustar anchos
ws_resumen.column_dimensions['A'].width = 35
ws_resumen.column_dimensions['B'].width = 15
ws_resumen.column_dimensions['C'].width = 15

# ----- HOJA 2: TABLA CRUZADA ESTADO vs ACCIÓN -----
ws_cruzada = wb.create_sheet("Estado vs Acción")

ws_cruzada['A1'] = "TABLA CRUZADA: ESTADO DUSA vs ACCIÓN TOMADA"
ws_cruzada['A1'].font = Font(bold=True, size=14)
ws_cruzada.merge_cells('A1:F1')

pivot_data = pd.crosstab(df['Estado_Limpio'], df['Accion_Limpia'], margins=True, margins_name='Total')

for r_idx, row in enumerate(dataframe_to_rows(pivot_data, index=True, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_cruzada.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 3 or r_idx == 4:  # Headers
            cell.fill = header_fill
            cell.font = header_font
        if c_idx == 1 and r_idx > 4:  # Index column
            cell.font = Font(bold=True)

# Ajustar anchos
for col_idx in range(1, 10):
    ws_cruzada.column_dimensions[chr(64 + col_idx)].width = 18

# ----- HOJA 3: PRODUCTOS DISPONIBLES -----
ws_disponibles = wb.create_sheet("✅ Disponibles")
df_disp = df[df['Estado_Limpio'] == 'Disponible'][['SKU', 'Título ML', 'Producto DUSA', 'Precio DUSA', 'Precio ML', 'Accion_Limpia']]
df_disp.columns = ['SKU', 'Título ML', 'Producto DUSA', 'Precio DUSA', 'Precio ML', 'Acción']

ws_disponibles['A1'] = f"PRODUCTOS DISPONIBLES EN DUSA ({len(df_disp)} productos)"
ws_disponibles['A1'].font = Font(bold=True, size=14, color="155724")
ws_disponibles.merge_cells('A1:F1')

for r_idx, row in enumerate(dataframe_to_rows(df_disp, index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_disponibles.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 3:
            cell.fill = header_fill
            cell.font = header_font

# ----- HOJA 4: PRODUCTOS FALTANTES -----
ws_faltantes = wb.create_sheet("❌ Faltantes")
df_falt = df[df['Estado_Limpio'] == 'Faltante'][['SKU', 'Título ML', 'Producto DUSA', 'Precio DUSA', 'Precio ML', 'Accion_Limpia']]
df_falt.columns = ['SKU', 'Título ML', 'Producto DUSA', 'Precio DUSA', 'Precio ML', 'Acción']

ws_faltantes['A1'] = f"PRODUCTOS FALTANTES EN DUSA ({len(df_falt)} productos)"
ws_faltantes['A1'].font = Font(bold=True, size=14, color="721c24")
ws_faltantes.merge_cells('A1:F1')

for r_idx, row in enumerate(dataframe_to_rows(df_falt, index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_faltantes.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 3:
            cell.fill = header_fill
            cell.font = header_font

# ----- HOJA 5: PRODUCTOS A CONSULTAR -----
ws_consultar = wb.create_sheet("⚠️ Consultar")
df_cons = df[df['Estado_Limpio'] == 'Consultar'][['SKU', 'Título ML', 'Producto DUSA', 'Precio DUSA', 'Precio ML', 'Accion_Limpia']]
df_cons.columns = ['SKU', 'Título ML', 'Producto DUSA', 'Precio DUSA', 'Precio ML', 'Acción']

ws_consultar['A1'] = f"PRODUCTOS A CONSULTAR ({len(df_cons)} productos)"
ws_consultar['A1'].font = Font(bold=True, size=14, color="856404")
ws_consultar.merge_cells('A1:F1')

for r_idx, row in enumerate(dataframe_to_rows(df_cons, index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_consultar.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 3:
            cell.fill = header_fill
            cell.font = header_font

# ----- HOJA 6: NO ENCONTRADOS -----
ws_no_enc = wb.create_sheet("🔍 No encontrados")
df_no = df[df['Estado_Limpio'] == 'No encontrado'][['SKU', 'Título ML', 'Precio ML', 'Stock ML', 'Accion_Limpia']]
df_no.columns = ['SKU', 'Título ML', 'Precio ML', 'Stock ML', 'Acción']

ws_no_enc['A1'] = f"PRODUCTOS NO ENCONTRADOS EN DUSA ({len(df_no)} productos)"
ws_no_enc['A1'].font = Font(bold=True, size=14, color="383d41")
ws_no_enc.merge_cells('A1:E1')

for r_idx, row in enumerate(dataframe_to_rows(df_no, index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_no_enc.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 3:
            cell.fill = header_fill
            cell.font = header_font

# ----- HOJA 7: PRODUCTOS CON ACCIONES -----
ws_acciones = wb.create_sheet("📋 Con Acciones")
df_acc = df[df['Accion_Limpia'] != 'Sin acción'][['SKU', 'Título ML', 'Estado_Limpio', 'Producto DUSA', 'Accion_Limpia']]
df_acc.columns = ['SKU', 'Título ML', 'Estado DUSA', 'Producto DUSA', 'Acción Tomada']

ws_acciones['A1'] = f"PRODUCTOS CON ACCIONES TOMADAS ({len(df_acc)} productos)"
ws_acciones['A1'].font = Font(bold=True, size=14, color="1e3c72")
ws_acciones.merge_cells('A1:E1')

for r_idx, row in enumerate(dataframe_to_rows(df_acc, index=False, header=True), 3):
    for c_idx, value in enumerate(row, 1):
        cell = ws_acciones.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if r_idx == 3:
            cell.fill = header_fill
            cell.font = header_font
        # Colorear según acción
        if c_idx == 5 and r_idx > 3:
            if value == 'Activado':
                cell.fill = disponible_fill
            elif value == 'No activado':
                cell.fill = faltante_fill
            elif value == 'Diferida':
                cell.fill = consultar_fill

# Ajustar anchos en todas las hojas
for ws in wb.worksheets:
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        if col == 'B':
            ws.column_dimensions[col].width = 50
        else:
            ws.column_dimensions[col].width = 18

# Guardar
archivo_excel = f'{CARPETA_SALIDA}/analisis_dusa_completo.xlsx'
wb.save(archivo_excel)
print(f"   ✅ Guardado: analisis_dusa_completo.xlsx")

# ============================================
# RESUMEN FINAL
# ============================================
print("\n" + "="*50)
print("✅ ANÁLISIS COMPLETADO")
print("="*50)
print(f"\n📁 Archivos guardados en: {CARPETA_SALIDA}")
print("\n📊 Gráficos generados:")
print("   • grafico_estados_dusa.png")
print("   • grafico_barras_estados.png")
print("   • grafico_acciones.png")
print("   • grafico_acciones_por_estado.png")
print("\n📑 Excel con tablas dinámicas:")
print("   • analisis_dusa_completo.xlsx")
print("     - Resumen General")
print("     - Estado vs Acción (tabla cruzada)")
print("     - ✅ Disponibles")
print("     - ❌ Faltantes")
print("     - ⚠️ Consultar")
print("     - 🔍 No encontrados")
print("     - 📋 Con Acciones")

print(f"\n📈 ESTADÍSTICAS:")
print(f"   Total productos: {total:,}")
print(f"   ✅ Disponibles: {estados_count.get('Disponible', 0):,}")
print(f"   ❌ Faltantes: {estados_count.get('Faltante', 0):,}")
print(f"   ⚠️ Consultar: {estados_count.get('Consultar', 0):,}")
print(f"   🔍 No encontrados: {estados_count.get('No encontrado', 0):,}")
print(f"\n   Acciones tomadas: {len(df_con_accion):,}")

# Abrir carpeta en Finder
import subprocess
subprocess.run(['open', CARPETA_SALIDA])
